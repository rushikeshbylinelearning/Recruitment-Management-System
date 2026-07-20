#!/usr/bin/env python3
"""
Generate the standard Byline HR bulk-upload Excel template.
Uses official stage colors from backend/config/stageColorMapping.js
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "Bulk_Upload_Template.xlsx"

# Official app colors (Candidate Name cell ONLY)
COLORS = {
    "header": "FF4472C4",
    "applied": None,
    "selected": "FF92D050",       # Shortlisted -> Selected column
    "hired": "FF00B050",          # Hired -> Hired column (NOT light blue)
    "rejected": "FFFF0000",
    "not_relevant": "FFFFC000",
    "follow_up": "FF00B0F0",
    "came_down": "FFFFFF00",
    "no_show": "FF7030A0",
    "no_response": "FF7F7F7F",
    "legend_bg": "FFF2F2F2",
    "instruction_bg": "FFFFF8E1",
}

HEADERS = [
    "Candidate Name",
    "Position",
    "Email",
    "Phone",
    "Stage",
    "Remarks",
    "Source",
]

SAMPLE_ROWS = [
    ("Example Applicant", "Full Stack Developer", "applicant@email.com", "9876543210", "", "New application", "Naukri"),
    ("Shortlisted Person", "Content Writer", "shortlisted@email.com", "9876543211", "Selected", "Offer discussion pending", "Indeed"),
    ("Hired Employee", "HR Executive", "hired@email.com", "9876543212", "Hired", "Joined on 15 Jan", "Referral"),
    ("Rejected Candidate", "Graphic Designer", "rejected@email.com", "9876543213", "Rejected", "Profile not matched", "LinkedIn"),
    ("Follow Up Person", "Digital Marketing Executive", "followup@email.com", "9876543214", "Follow Up", "Call back next week", "Manual Entry"),
    ("Interview Attended", "Instructional Designer", "interview@email.com", "9876543215", "Interview", "Came for interview", "Manual Entry"),
    ("Not Relevant Profile", "Animator", "nr@email.com", "9876543216", "Not Relevant", "Skills mismatch", "Manual Entry"),
    ("No Response", "Project Manager/Coordinator", "noresponse@email.com", "9876543217", "Follow Up", "No response to calls", "Manual Entry"),
]

NAME_FILLS = [
    None,
    COLORS["selected"],
    COLORS["hired"],
    COLORS["rejected"],
    COLORS["follow_up"],
    COLORS["came_down"],
    COLORS["not_relevant"],
    COLORS["no_response"],
]

POSITION_ALIASES = [
    ("Use in Position column", "Maps to job card / posting"),
    ("FSD, Full Stack Developer", "Full Stack Developer"),
    ("GD, Graphic Designer", "Graphic Designer"),
    ("ID, Instructional Designer", "Instructional Designers"),
    ("Content, Content Writer", "Content Writers"),
    ("BDE, Sales, Marketing Executive", "Sales and Marketing"),
    ("DM, Digital Marketing Executive", "Digital Marketing"),
    ("HR, HR Executive, HR Intern", "Human Resource"),
    ("PC, Project Coordinator", "Project Coordinators"),
    ("Animator, 2D Animator", "Animators"),
    ("ELD, E-Learning Developer", "IT"),
]


def fill(hex_argb: str | None) -> PatternFill | None:
    if not hex_argb:
        return None
    return PatternFill(fill_type="solid", fgColor=hex_argb)


def thin_border() -> Border:
    s = Side(style="thin", color="FFD9D9D9")
    return Border(left=s, right=s, top=s, bottom=s)


def style_header(ws, row: int, headers: list[str]) -> None:
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.font = Font(bold=True, color="FFFFFFFF", size=11)
        cell.fill = fill(COLORS["header"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = max(16, len(title) + 4)


def build_import_sheet(wb: openpyxl.Workbook) -> None:
    ws = wb.active
    ws.title = "Import Data"
    style_header(ws, 1, HEADERS)

    for idx, (row_data, name_color) in enumerate(zip(SAMPLE_ROWS, NAME_FILLS), start=2):
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=idx, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border()
            if col == 1 and name_color:
                cell.fill = fill(name_color)
        ws.row_dimensions[idx].height = 28

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(SAMPLE_ROWS)+1}"


def build_instructions_sheet(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Instructions")
    ws.column_dimensions["A"].width = 110

    lines = [
        "BYLINE HR — BULK UPLOAD EXCEL TEMPLATE",
        "",
        "HOW TO USE",
        "1. Fill data in the 'Import Data' sheet (replace sample rows with real candidates).",
        "2. Keep the header row exactly as-is — the app auto-maps these column names.",
        "3. Upload from Jobs page -> Applicants -> Bulk Import (or Candidates -> Import).",
        "4. One row = one candidate. Use the Position column to route to the correct job card.",
        "",
        "STAGE / STATUS (MOST IMPORTANT)",
        "• Color ONLY the Candidate Name cell — not the full row.",
        "• The app reads stage from the Candidate Name cell background color.",
        "• You may also type in the Stage column as backup text.",
        "",
        "OFFICIAL COLORS (Candidate Name cell)",
        "  Light Green  #92D050  ->  Selected (shortlisted, NOT hired)",
        "  Dark Green   #00B050  ->  Hired",
        "  Red          #FF0000  ->  Rejected",
        "  Gold         #FFC000  ->  Not Relevant",
        "  Blue         #00B0F0  ->  Follow Up",
        "  Yellow       #FFFF00  ->  Came Down for Interview",
        "  Purple       #7030A0  ->  Didn't Come for Interview",
        "  Gray         #7F7F7F  ->  Didn't Respond",
        "  No fill      white    ->  Applied (default)",
        "",
        "DO NOT USE",
        "  • Light blue #9BC2E6 for hired — the app will NOT detect this.",
        "  • Coloring entire rows — supported (legacy Book1-style trackers).",
        "    Row color is used when the Candidate Name cell has no fill.",
        "  • Mixed roles without a Position value on each row.",
        "",
        "POSITION COLUMN",
        "  Use full titles (Full Stack Developer) or aliases (FSD, GD, ID, PC, BDE).",
        "  See 'Position Guide' sheet for common abbreviations.",
        "",
        "OPTIONAL COLUMNS",
        "  Email, Phone, Remarks, Source — Remarks is saved as HR notes.",
        "",
        "DELETE sample rows before uploading production data, or the app will import them too.",
    ]

    for i, line in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if i == 1:
            cell.font = Font(bold=True, size=14, color="FF1F4E79")
        elif line.endswith(":") or line.isupper() and line:
            cell.font = Font(bold=True, size=11)


def build_color_legend_sheet(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Color Legend")
    headers = ["Stage", "Kanban column", "Candidate Name color", "Stage text (optional)"]
    style_header(ws, 1, headers)

    rows = [
        ("Applied", "Applied", None, "(leave blank)"),
        ("Selected / Shortlisted", "Selected", COLORS["selected"], "Selected"),
        ("Hired", "Hired", COLORS["hired"], "Hired"),
        ("Rejected", "Rejected", COLORS["rejected"], "Rejected"),
        ("Not Relevant", "Rejected (Not Relevant)", COLORS["not_relevant"], "Not Relevant"),
        ("Follow Up", "Follow Up", COLORS["follow_up"], "Follow Up"),
        ("Came Down for Interview", "Interview", COLORS["came_down"], "Interview"),
        ("Didn't Come for Interview", "Interview (No Show)", COLORS["no_show"], "No Show"),
        ("Didn't Respond", "Follow Up (No Response)", COLORS["no_response"], "Follow Up"),
    ]

    for r, (stage, column, color, text) in enumerate(rows, start=2):
        ws.cell(r, 1, stage).border = thin_border()
        ws.cell(r, 2, column).border = thin_border()
        swatch = ws.cell(r, 3, "  Sample  ")
        swatch.border = thin_border()
        swatch.alignment = Alignment(horizontal="center")
        if color:
            swatch.fill = fill(color)
        ws.cell(r, 4, text).border = thin_border()

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 22


def build_position_guide_sheet(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Position Guide")
    style_header(ws, 1, ["Excel Position value", "Job card / posting"])
    for r, (alias, card) in enumerate(POSITION_ALIASES, start=2):
        ws.cell(r, 1, alias).border = thin_border()
        ws.cell(r, 2, card).border = thin_border()
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 32


def main() -> None:
    wb = openpyxl.Workbook()
    build_import_sheet(wb)
    build_instructions_sheet(wb)
    build_color_legend_sheet(wb)
    build_position_guide_sheet(wb)
    wb.save(OUT)
    print(f"Created: {OUT}")


if __name__ == "__main__":
    main()
