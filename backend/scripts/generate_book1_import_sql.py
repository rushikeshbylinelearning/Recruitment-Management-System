#!/usr/bin/env python3
"""
Generate SQL import script from Book1.xlsx (Khushi tracker sheet).
Run: python backend/scripts/generate_book1_import_sql.py
Output: backend/migrations/import_book1_candidates.sql
"""

from __future__ import annotations

import json
import re
import uuid
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "Book1.xlsx"
OUT = ROOT / "backend" / "migrations" / "import_book1_candidates.sql"

# Excel row colors (Book1 uses non-standard shades vs bulk-upload defaults)
COLOR_HIRED = "#9BC2E6"
COLOR_SELECTED = "#92D050"
COLOR_REJECTED = "#FF0000"

# Excel alias -> (job_id, canonical position for job-card category matching)
POSITION_MAP: dict[str, tuple[int, str]] = {
    "digital marketing": (25, "Digital Marketing Executive"),
    "digital marketing executive": (25, "Digital Marketing Executive"),
    "digital marketing": (25, "Digital Marketing Executive"),
    "digital m arketing": (25, "Digital Marketing Executive"),
    "dm": (25, "Digital Marketing Executive"),
    "bde": (11, "Business Development Executive"),
    "fsd": (19, "Full Stack Developer"),
    "senior fsd": (8, "Senior - Full Stack Developer"),
    "fsc": (19, "Full Stack Developer"),
    "content": (7, "Content Writer"),
    "content writer": (7, "Content Writer"),
    "content writers": (7, "Content Writer"),
    "content/ id": (16, "Instructional Designer"),
    "content/id": (16, "Instructional Designer"),
    "id": (16, "Instructional Designer"),
    "icw": (17, "Instructional Content Writer"),
    "hr intern": (28, "HR Executive"),
    "hr inetrn": (28, "HR Executive"),
    "hr": (28, "HR Executive"),
    "intern": (18, "Intern & Fresher Sales/Marketing"),
    "intern sales and marketing": (18, "Intern & Fresher Sales/Marketing"),
    "sales and marketing": (11, "Sales and Marketing"),
    "project coordinator": (26, "Project Manager/Coordinator"),
    "pc": (26, "Project Manager/Coordinator"),
    "marketing": (11, "Marketing Executive"),
    "marketing executive": (11, "Marketing Executive"),
    "markering": (11, "Marketing Executive"),
    "sales": (30, "Sales"),
    "animator": (23, "Animator"),
    "2d animator": (23, "Animator"),
    "gd": (10, "Graphic Designer"),
    "ui/ux": (10, "Graphic Designer"),
    "visualizer": (14, "Visualizer"),
    "eld": (4, "E-Learning Developer"),
    "e-learning developer": (4, "E-Learning Developer"),
    "product manager": (26, "Project Manager/Coordinator"),
    "product  manager": (26, "Project Manager/Coordinator"),
    "it": (27, "Senior IT"),
    "senior it": (27, "Senior IT"),
}

REJECT_KEYWORDS = (
    "reject",
    "not selected",
    "not a relevant",
    "not relevant",
    "lack of",
    "ditched",
    "did not show",
    "didn't show",
    "did not attend",
    "didn't attend",
    "did not come",
    "didn't come",
    "no show",
    "fake experience",
    "out of budget",
    "language issue",
    "attitude issue",
    "not a graduate",
    "not an id",
    "not core id",
    "profile not matched",
    "on hold by",
)


def sql_escape(value: str | None) -> str:
    if value is None:
        return "NULL"
    return "'" + str(value).replace("\\", "\\\\").replace("'", "''") + "'"


def get_cell_color(cell) -> str | None:
    fill = cell.fill
    if not fill or not fill.fgColor:
        return None
    rgb = fill.fgColor.rgb
    if not rgb or str(rgb) in ("00000000", "FFFFFFFF", "None"):
        return None
    s = str(rgb)
    if len(s) == 8:
        return "#" + s[2:].upper()
    return "#" + s.upper()


def normalize_position_key(raw: str | None) -> str:
    if not raw:
        return ""
    return re.sub(r"\s+", " ", str(raw).strip().lower())


def map_position(raw: str | None) -> tuple[int | None, str]:
    key = normalize_position_key(raw)
    if not key:
        return None, "Unknown"
    if key in POSITION_MAP:
        job_id, canonical = POSITION_MAP[key]
        return job_id, canonical
    # partial / fuzzy fallbacks
    for alias, (job_id, canonical) in sorted(POSITION_MAP.items(), key=lambda x: -len(x[0])):
        if alias in key or key in alias:
            return job_id, canonical
    return None, str(raw).strip().title()


def parse_applied_date(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s[:19], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def detect_stage(row_colors: list[str | None], status: str | None, remarks: str | None) -> tuple[str, str, str | None]:
    """Return (legacy_stage, main_stage, sub_stage)."""
    colors = {c for c in row_colors if c}
    status_norm = (status or "").strip().lower()
    remarks_norm = (remarks or "").strip().lower()

    if COLOR_HIRED in colors or status_norm == "hired":
        return "Hired", "hired", None
    if COLOR_SELECTED in colors or status_norm == "selected":
        return "Selected", "selected", None
    if COLOR_REJECTED in colors or status_norm == "rejected":
        return "Rejected", "rejected", "rejected"

    if any(k in remarks_norm for k in REJECT_KEYWORDS):
        return "Rejected", "rejected", "rejected"

    if "on hold" in remarks_norm and "reject" not in remarks_norm:
        return "On Hold", "rejected", "on-hold"

    return "Applied", "applied", None


def append_mysql_add_column_if_missing(lines: list[str], table: str, column: str, definition: str) -> None:
    """Emit idempotent ADD COLUMN compatible with MySQL 5.7 / MariaDB (no IF NOT EXISTS)."""
    lines.append(
        f"SET @sql = IF("
        f"(SELECT COUNT(*) FROM information_schema.COLUMNS "
        f"WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = '{table}' AND COLUMN_NAME = '{column}') = 0, "
        f"'ALTER TABLE `{table}` ADD COLUMN `{column}` {definition}', "
        f"'SELECT 1');"
    )
    lines.append("PREPARE stmt FROM @sql;")
    lines.append("EXECUTE stmt;")
    lines.append("DEALLOCATE PREPARE stmt;")
    lines.append("")


def append_prerequisite_sql(lines: list[str]) -> None:
    lines.append("-- Prerequisites (MySQL 5.7 / MariaDB compatible — skips existing columns)")
    append_mysql_add_column_if_missing(lines, "candidates", "main_stage", "VARCHAR(50) NULL AFTER stage")
    append_mysql_add_column_if_missing(lines, "candidates", "sub_stage", "VARCHAR(50) NULL AFTER main_stage")
    append_mysql_add_column_if_missing(lines, "candidates", "uploaded_by", "INT(11) NULL AFTER assigned_to")
    lines.append("ALTER TABLE candidates")
    lines.append("  MODIFY COLUMN stage ENUM(")
    lines.append("    'Applied','Follow Up','Screening','Interview','Offer','Hired',")
    lines.append("    'On Hold','Rejected','No Show - Interview','No Show - Onboarding',")
    lines.append("    'Last Minute Back Out','Profile Not Matched','Selected'")
    lines.append("  ) NOT NULL DEFAULT 'Applied';")
    lines.append("")
    lines.append("ALTER TABLE hr_notes")
    lines.append("  MODIFY COLUMN stage ENUM(")
    lines.append("    'Applied','Follow Up','Screening','Interview','Offer','Hired',")
    lines.append("    'On Hold','Rejected','No Show - Interview','No Show - Onboarding',")
    lines.append("    'Last Minute Back Out','Profile Not Matched','Selected'")
    lines.append("  ) NOT NULL DEFAULT 'Applied';")
    lines.append("")


def build_notes(remarks, hr_comment) -> str | None:
    parts = []
    if remarks and str(remarks).strip():
        parts.append(str(remarks).strip())
    if hr_comment and str(hr_comment).strip():
        parts.append(f"HR Comment: {str(hr_comment).strip()}")
    if not parts:
        return None
    return " | ".join(parts)


def main() -> None:
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active

    raw_rows = []
    unmapped_positions: Counter[str] = Counter()

    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 2).value
        if not name or not str(name).strip():
            continue

        raw_position = ws.cell(r, 3).value
        status = ws.cell(r, 4).value
        remarks = ws.cell(r, 5).value
        hr_comment = ws.cell(r, 6).value
        applied = parse_applied_date(ws.cell(r, 1).value)

        row_colors = [get_cell_color(ws.cell(r, c)) for c in range(1, 7)]
        has_color = any(row_colors)
        position_text = str(raw_position or "").strip()
        remarks_text = str(remarks or "").strip()
        status_text = str(status or "").strip()

        # Skip blank duplicate stubs (same person listed again on next row with full data)
        if not position_text and not remarks_text and not status_text and not has_color:
            continue

        # Infer position from remarks when the Position cell is empty
        infer_source = f"{position_text} {remarks_text} {status_text}".lower()
        if not position_text:
            if "project manager" in infer_source or "project coordinator" in infer_source:
                position_text = "Project Coordinator"
            elif "sales" in infer_source or "marketing" in infer_source:
                position_text = "Sales and Marketing"

        legacy_stage, main_stage, sub_stage = detect_stage(row_colors, status, remarks)
        job_id, canonical_position = map_position(position_text)
        if job_id is None and position_text:
            unmapped_positions[position_text] += 1

        notes = build_notes(remarks, hr_comment)
        completeness = (
            (10 if position_text else 0)
            + (5 if remarks_text else 0)
            + (3 if status_text else 0)
            + (2 if applied else 0)
            + (5 if has_color else 0)
            + (2 if notes else 0)
        )

        raw_rows.append(
            {
                "excel_row": r,
                "name_key": str(name).strip().lower(),
                "name": str(name).strip()[:100],
                "raw_position": position_text,
                "position": canonical_position[:200],
                "job_id": job_id,
                "applied_date": applied,
                "stage": legacy_stage,
                "main_stage": main_stage,
                "sub_stage": sub_stage,
                "notes": notes,
                "completeness": completeness,
            }
        )

    # Deduplicate by candidate name — keep the most complete row
    best_by_name: dict[str, dict] = {}
    for row in raw_rows:
        key = row["name_key"]
        existing = best_by_name.get(key)
        if existing is None or row["completeness"] > existing["completeness"] or (
            row["completeness"] == existing["completeness"] and row["excel_row"] > existing["excel_row"]
        ):
            best_by_name[key] = row

    rows = []
    for row in best_by_name.values():
        row["id"] = str(uuid.uuid4())
        rows.append(row)
    rows.sort(key=lambda x: x["excel_row"])

    stage_counts = Counter(r["stage"] for r in rows)
    job_counts = Counter(r["job_id"] for r in rows if r["job_id"])
    card_counts: dict[str, int] = defaultdict(int)
    card_aliases = {
        "Sales and Marketing": {"sales", "business development executive", "marketing executive", "intern & fresher sales/marketing", "linkedin profile sales"},
        "Animators": {"animator"},
        "Graphic Designer": {"graphic designer", "visualizer"},
        "Full Stack Developer": {"full stack developer", "senior - full stack developer"},
        "Content Writers": {"content writer", "instructional content writer"},
        "Instructional Designers": {"instructional designer"},
        "Digital Marketing": {"digital marketing executive"},
        "Human Resource": {"hr executive"},
        "IT": {"e-learning developer", "senior it"},
        "Project Coordinators": {"project manager/coordinator"},
    }
    alias_to_card = {}
    for card, aliases in card_aliases.items():
        for a in aliases:
            alias_to_card[a.lower()] = card
    for row in rows:
        card = alias_to_card.get(row["position"].lower())
        if card:
            card_counts[card] += 1

    lines: list[str] = []
    lines.append("-- ============================================================")
    lines.append("-- Import Book1.xlsx candidate tracker (312 unique candidates, 322 Excel rows)")
    lines.append("-- Uploaded by: Khushi (users.username = 'khushi')")
    lines.append("-- Source file: Book1.xlsx")
    lines.append("--")
    lines.append("-- Color mapping (row-level, any column):")
    lines.append(f"--   Blue   {COLOR_HIRED} -> Hired")
    lines.append(f"--   Green  {COLOR_SELECTED} -> Selected (shortlisted, NOT hired)")
    lines.append(f"--   Red    {COLOR_REJECTED} -> Rejected (none in file; inferred from Status/Remarks)")
    lines.append("--")
    lines.append("-- Generated by: backend/scripts/generate_book1_import_sql.py")
    lines.append("-- ============================================================")
    lines.append("")
    lines.append("SET NAMES utf8mb4;")
    lines.append("SET @import_started_at = NOW();")
    lines.append("SET @khushi_user_id = (SELECT id FROM users WHERE username = 'khushi' LIMIT 1);")
    lines.append("")
    append_prerequisite_sql(lines)
    lines.append("SELECT IF(@khushi_user_id IS NULL,")
    lines.append("  (SELECT 'ERROR: Khushi user not found. Run backend/migrations/add_recruiter_khushi.sql first.' AS message),")
    lines.append("  (SELECT CONCAT('Khushi user id = ', @khushi_user_id) AS message));")
    lines.append("")
    lines.append("-- Optional safety backup")
    lines.append("CREATE TABLE IF NOT EXISTS candidates_backup_book1_import AS")
    lines.append("SELECT * FROM candidates WHERE 1 = 0;")
    lines.append("")
    lines.append("START TRANSACTION;")
    lines.append("")

    # import log
    lines.append("INSERT INTO import_logs (user_id, filename, total_rows, success_count, failure_count, processing_time, uploaded_at)")
    lines.append(
        f"VALUES (@khushi_user_id, 'Book1.xlsx', {len(rows)}, {len(rows)}, 0, 0, @import_started_at);"
    )
    lines.append("SET @import_log_id = LAST_INSERT_ID();")
    lines.append("")

    for row in rows:
        applied_sql = sql_escape(row["applied_date"]) if row["applied_date"] else "CURDATE()"
        job_sql = str(row["job_id"]) if row["job_id"] is not None else "NULL"
        sub_stage_sql = sql_escape(row["sub_stage"]) if row["sub_stage"] else "NULL"
        notes_sql = sql_escape(row["notes"]) if row["notes"] else "NULL"

        lines.append(
            "INSERT INTO candidates ("
            "id, job_id, name, email, phone, position, stage, main_stage, sub_stage, "
            "source, applied_date, notes, uploaded_by, in_house_assignment_status, skills"
            ") VALUES ("
            f"{sql_escape(row['id'])}, "
            f"{job_sql}, "
            f"{sql_escape(row['name'])}, "
            "NULL, NULL, "
            f"{sql_escape(row['position'])}, "
            f"{sql_escape(row['stage'])}, "
            f"{sql_escape(row['main_stage'])}, "
            f"{sub_stage_sql}, "
            "'Manual Entry', "
            f"{applied_sql}, "
            f"{notes_sql}, "
            "@khushi_user_id, "
            "'Pending', "
            "'[]'"
            ");"
        )

        if row["notes"]:
            lines.append(
                "INSERT INTO hr_notes (candidate_id, stage, note_text, interaction_type, author_id) VALUES ("
                f"{sql_escape(row['id'])}, "
                f"{sql_escape(row['stage'])}, "
                f"{notes_sql}, "
                "'General Note', "
                "@khushi_user_id"
                ");"
            )
        lines.append("")

    # refresh applicant counts for touched jobs
    touched_jobs = sorted(j for j in job_counts if j is not None)
    lines.append("-- Refresh applicant_count on affected job postings")
    for jid in touched_jobs:
        lines.append(
            "UPDATE job_postings SET applicant_count = ("
            f"SELECT COUNT(*) FROM candidates WHERE job_id = {jid}"
            f"), updated_at = NOW() WHERE id = {jid};"
        )

    lines.append("")
    lines.append("COMMIT;")
    lines.append("")
    lines.append("-- ============================================================")
    lines.append("-- Verification queries")
    lines.append("-- ============================================================")
    lines.append("SELECT stage, COUNT(*) AS cnt FROM candidates")
    lines.append("WHERE uploaded_by = @khushi_user_id AND source = 'Manual Entry'")
    lines.append("  AND created_at >= @import_started_at")
    lines.append("GROUP BY stage ORDER BY cnt DESC;")
    lines.append("")
    lines.append("SELECT j.id, j.title, COUNT(c.id) AS imported_count")
    lines.append("FROM candidates c")
    lines.append("JOIN job_postings j ON j.id = c.job_id")
    lines.append("WHERE c.uploaded_by = @khushi_user_id")
    lines.append("  AND c.created_at >= @import_started_at")
    lines.append("GROUP BY j.id, j.title")
    lines.append("ORDER BY imported_count DESC;")
    lines.append("")
    lines.append("-- Expected stage distribution from Excel:")
    for stage, cnt in stage_counts.most_common():
        lines.append(f"--   {stage}: {cnt}")
    lines.append("-- Expected job-card distribution (by position alias):")
    for card, cnt in sorted(card_counts.items(), key=lambda x: -x[1]):
        lines.append(f"--   {card}: {cnt}")
    if unmapped_positions:
        lines.append("-- UNMAPPED positions (job_id NULL):")
        for pos, cnt in unmapped_positions.most_common():
            lines.append(f"--   {pos!r}: {cnt}")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text("\n".join(lines) + "\n", encoding="utf-8")

    summary = {
        "rows": len(rows),
        "stages": dict(stage_counts),
        "jobs": {str(k): v for k, v in job_counts.items()},
        "cards": dict(card_counts),
        "unmapped": dict(unmapped_positions),
        "output": str(OUT),
    }
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
